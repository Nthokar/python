import csv
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
import re

from openpyxl.descriptors import (
    String,
    Sequence,
    Integer,
)
from openpyxl.descriptors.serialisable import Serialisable

currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}

def сsv_parser(file_name):
    with open(file_name, encoding='utf_8_sig') as r_file:
        file_reader = csv.reader(r_file, delimiter=",")
        list_naming = file_reader.__next__()
        vacancies, vacancies_city = {}, {}
        for row in file_reader:
            if len(row) != len(list_naming) or row.__contains__(""):
                continue
            fieds = {}
            for field_index in range(len(row)):
                fieds.update({list_naming[field_index]: row[field_index]})
            year = fieds["published_at"][:4]
            vacancy = Vacancy(name=fieds["name"], salary_from=fieds["salary_from"], salary_to=fieds["salary_to"], salary_currency=fieds["salary_currency"], area_name=fieds["area_name"], published_at=fieds["published_at"])
            #vacancy2nd = Vacancy(row=row)
            if vacancies_city.keys().__contains__(vacancy.area_name):
                vacancies_city[vacancy.area_name].append(vacancy)
            else:
                vacancies_city.update({vacancy.area_name: [vacancy]})
            if vacancies.keys().__contains__(year):
                vacancies[year].append(vacancy)
            else:
                vacancies.update({year: [vacancy]})

        return vacancies, vacancies_city

filename, name = input("Введите название файла: "), input("Введите название профессии: ")

class Vacancy:
    def __init__(self, name, salary_from, salary_to, salary_currency, area_name, published_at):
        self.name = name
        self.salary_from = float(salary_from)
        self.salary_to = float(salary_to)
        self.salary_average = (self.salary_from + self.salary_to) / 2
        self.salary_currency = salary_currency
        self.area_name = area_name
        self.published_at = published_at

class Report:
    font_size = 11

    def generate_excel(self, statics_by_years, statics_by_cities):
        wb = Workbook()
        wb.remove(wb['Sheet'])

        statistics_by_year_sheet = wb.create_sheet("Статистика по годам")
        statistics_by_year_sheet.append(["Год", "Средняя зарплата", f"Средняя зарплата - {name}", "Количество вакансий", f"Количество вакансий - {name}"])
        for key in statics_by_years:
            statistics_by_year_sheet.append([key] + statics_by_years[key])

        cols_dict = {}
        for row in statistics_by_year_sheet.rows:
            for cell in row:
                letter = cell.column_letter
                cell.border = Border(left=Side(style='thin'),
                                     right=Side(style='thin'),
                                     top=Side(style='thin'),
                                     bottom=Side(style='thin'))
                if cell.value:
                    cell.font = Font(name='Calibri', size=self.font_size)
                    len_cell = len(str(cell.value))
                    len_cell_dict = 0
                    if letter in cols_dict:
                        len_cell_dict = cols_dict[letter]

                    if len_cell > len_cell_dict:
                        cols_dict[letter] = len_cell
                        ###!!! ПРОБЛЕМА АВТОМАТИЧЕСКОЙ ПОДГОНКИ !!!###
                        ###!!! расчет новой ширины колонки (здесь надо подгонять) !!!###
                        new_width_col = len_cell * self.font_size ** (self.font_size * 0.009)
                        statistics_by_year_sheet.column_dimensions[cell.column_letter].width = new_width_col

        for cell in statistics_by_year_sheet[1:1]:
            cell.font = Font(name='Calibri', size=self.font_size, bold=True)


        statistics_by_cities_sheet = wb.create_sheet("Статистика по городам")
        statistics_by_cities_sheet.append(['Город', 'Уровень зарплат', ' ', 'Город', 'Доля вакансий'])
        for key in statics_by_cities:
            statistics_by_cities_sheet.append([key, statics_by_cities[key][0], ' ', key, statics_by_cities[key][1]])

        cols_dict = {}
        for row in statistics_by_cities_sheet.rows:
            for cell in row:
                letter = cell.column_letter
                cell.border = Border(left=Side(style='thin'),
                                     right=Side(style='thin'),
                                     top=Side(style='thin'),
                                     bottom=Side(style='thin'))
                if cell.value:
                    cell.font = Font(name='Calibri', size=self.font_size)
                    len_cell = len(str(cell.value))
                    len_cell_dict = 0
                    if letter in cols_dict:
                        len_cell_dict = cols_dict[letter]

                    if len_cell > len_cell_dict:
                        cols_dict[letter] = len_cell
                        ###!!! ПРОБЛЕМА АВТОМАТИЧЕСКОЙ ПОДГОНКИ !!!###
                        ###!!! расчет новой ширины колонки (здесь надо подгонять) !!!###
                        new_width_col = len_cell * self.font_size ** (self.font_size * 0.009)
                        statistics_by_cities_sheet.column_dimensions[cell.column_letter].width = new_width_col


        for cell in statistics_by_cities_sheet['E']:
            cell.number_format = '0.00%'
        for cell in statistics_by_cities_sheet[1:1]:
            cell.font = Font(name='Calibri', size=self.font_size, bold=True)
        wb.save('report.xlsx')







vacancies, vacancies_city = сsv_parser(filename)

if (name == ""):
    vacancy_name = name
else:
    vacancy_name = ""


vacancies_count = sum(len(vacancies_city[i]) for i in vacancies_city)
city_to_pop = []
for city in vacancies_city:
    if (len(vacancies_city[city]) < int(vacancies_count / 100)):
        city_to_pop.append(city)

for city in city_to_pop:
    vacancies_city.pop(city)

map(lambda k, v: (k, v) if len(vacancies_city[k]) > vacancies_count / 100 else None, vacancies_city)

#print(vacancies_count)
#print(vacancies_city)

statistics_by_years = {}
vacancies_count_by_years = {}
vacancies_salary_by_years = {}
vacancies_count_by_years_for_name = {}
vacancies_salary_by_years_for_name = {}
for key in vacancies.keys():
     vacancies_count_by_years.update({key: len(vacancies[key])})
     vacancies_salary_by_years.update({key: int(sum(x.salary_average * currency_to_rub[x.salary_currency] for x in vacancies[key]) / vacancies_count_by_years[key])})
     vacancies_count_by_years_for_name.update({key: len(list(filter(lambda x: (name in x.name), vacancies[key])))})
     if vacancies_count_by_years_for_name[key] == 0:
         vacancies_salary_by_years_for_name.update({key: 0})
     else:
        vacancies_salary_by_years_for_name.update({key: int(sum(x.salary_average * currency_to_rub[x.salary_currency] for x in list(filter(lambda x: (name in x.name), vacancies[key]))) / vacancies_count_by_years_for_name[key])})
     statistics_by_years.update({key: [vacancies_salary_by_years[key], vacancies_salary_by_years_for_name[key], vacancies_count_by_years[key], vacancies_count_by_years_for_name[key]]})

statistics_by_cities = {}
vacancies_salary_by_city, vacancies_proportion_by_city = {}, {}
for key in vacancies_city.keys():
    vacancies_salary_by_city.update({key: int(sum(map(lambda x: x.salary_average * currency_to_rub[x.salary_currency], vacancies_city[key])) / len(
        vacancies_city[key]))})
    vacancies_proportion_by_city.update({key: float(len(vacancies_city[key])/vacancies_count)})
    statistics_by_cities.update({key: [vacancies_salary_by_city[key], vacancies_proportion_by_city[key]]})

Report().generate_excel(statistics_by_years, statistics_by_cities)

temp = '\''
print(f"Динамика уровня зарплат по годам: {str(vacancies_salary_by_years).replace(temp, '')}")
print("Динамика количества вакансий по годам: " + str(vacancies_count_by_years).replace(temp, ''))
print("Динамика уровня зарплат по годам для выбранной профессии: " + str(vacancies_salary_by_years_for_name).replace(temp, ''))
print("Динамика количества вакансий по годам для выбранной профессии: " + str(vacancies_count_by_years_for_name).replace(temp, ''))
vacancies_salary_by_city = {k: v for k, v in sorted(vacancies_salary_by_city.items(), key=lambda item: item[1], reverse=True)[0:10]}
vacancies_proportion_by_city = {k: round(v, 4) for k, v in sorted(vacancies_proportion_by_city.items(), key=lambda item: item[1], reverse=True)[0:10]}

#vacancies_salary_by_city = list(map(lambda k: (k, vacancies_salary_by_city[k]), vacancies_salary_by_city))
#vacancies_salary_by_city.sort(key=lambda tup: tup[1], reverse=True)
print("Уровень зарплат по городам (в порядке убывания): " + str(vacancies_salary_by_city))
print(f"Доля вакансий по городам (в порядке убывания): {str(vacancies_proportion_by_city)}")

